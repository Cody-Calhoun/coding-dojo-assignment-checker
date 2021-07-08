import openpyxl
import random
import math
import os
# This is declaring the path of where to look for your excel file
os.chdir('/Users/williamcalhoun/Desktop/')

# 
workbook = openpyxl.load_workbook('student_list.xlsx') #Need to Change this

sheet = workbook['Jun 2021 Will (New)'] # Need proper sheet name

students = []
weekly_students = []

core_assignments = {
    'week_one' : ['Hello World', 'Functions Basic 1', 'Functions Intermediate 2'],
    'week_two' : ['Bank Account', 'Users with Bank Accounts', 'Dojos and Ninjas'],
    'week_three': ['Time Display', 'Dojo Survey Revisited'],
    'week_four': ['Users with Templates'],
    'week_five': ['Books/Authors (Shell)'],
    'week_six': ['Semi-Restful TV Shows Validated', 'Login and Registration', 'Favorite Books'],
}

def load_students():
    for i in range(4, 45):
        student = sheet.cell(row=i, column=1).value + ' ' + sheet.cell(row=i, column=2).value
        students.append(student)
    return students

def get_assignments(week):
    for i in range(7):
        number = math.floor(random.randint(0, (len(students)-1)))
        weekly_students.append(f"{students[number]}: {random.choice(core_assignments[week])}")
        del students[number]
    print(weekly_students)
    return weekly_students

load_students()
get_assignments('week_two')

    